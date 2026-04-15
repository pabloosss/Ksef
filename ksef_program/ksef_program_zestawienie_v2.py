import os
import re
import sys
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright


TARGET_COLUMNS = [
    "Identyfikator sprzedawcy",
    "Nazwa sprzedawcy",
    "Nr KSeF",
    "Nr faktury",
    "Data wystawienia",
    "Data zapisania w KSeF",
    "Data otrzymania",
    "Waluta",
    "Netto",
    "Brutto",
    "VAT (PLN)",
]

HEADER_ALIASES = {
    "Identyfikator sprzedawcy": [
        "identyfikator sprzedawcy", "nip sprzedawcy", "nip", "sprzedawca nip"
    ],
    "Nazwa sprzedawcy": [
        "nazwa sprzedawcy", "sprzedawca", "nazwa podmiotu", "nazwa"
    ],
    "Nr KSeF": [
        "nr ksef", "numer ksef", "ksef", "ksef reference", "numer referencyjny ksef"
    ],
    "Nr faktury": [
        "nr faktury", "numer faktury", "faktura", "invoice number"
    ],
    "Data wystawienia": [
        "data wystawienia", "wystawiono", "invoice date"
    ],
    "Data zapisania w KSeF": [
        "data zapisania w ksef", "data przyjęcia w ksef", "data nadania numeru ksef", "zapisano w ksef"
    ],
    "Data otrzymania": [
        "data otrzymania", "otrzymano", "received date"
    ],
    "Waluta": [
        "waluta", "currency"
    ],
    "Netto": [
        "netto", "wartość netto", "kwota netto"
    ],
    "Brutto": [
        "brutto", "wartość brutto", "kwota brutto"
    ],
    "VAT (PLN)": [
        "vat (pln)", "vat pln", "vat", "kwota vat pln", "podatek vat pln"
    ],
}


class KsefSummaryApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("KSeF – Zestawienie FV")
        self.root.geometry("1220x820")
        self.root.minsize(1100, 760)
        self.root.configure(bg="#edf2f7")

        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None

        self.base_dir = os.getcwd()
        self.output_dir = os.path.join(self.base_dir, "zestawienia_ksef")
        os.makedirs(self.output_dir, exist_ok=True)

        today = datetime.today()
        first_day = today.replace(day=1)

        self.date_from_var = tk.StringVar(value=first_day.strftime("%Y-%m-%d"))
        self.date_to_var = tk.StringVar(value=today.strftime("%Y-%m-%d"))
        self.status_var = tk.StringVar(value="Status: gotowe")
        self.result_var = tk.StringVar(value="Wynik: nie sprawdzono")
        self.rows_var = tk.StringVar(value="0")
        self.pages_var = tk.StringVar(value="0")
        self.file_var = tk.StringVar(value="-")

        self.last_headers: List[str] = []
        self.last_rows: List[Dict] = []
        self.last_pages = 0

        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("Primary.TButton", font=("Segoe UI", 10, "bold"), padding=9)
        self.style.configure("Secondary.TButton", font=("Segoe UI", 10, "bold"), padding=9)
        self.style.configure("Danger.TButton", font=("Segoe UI", 10, "bold"), padding=9)
        self.style.configure(
            "Modern.Horizontal.TProgressbar",
            troughcolor="#d9e2ec",
            background="#d90429",
            bordercolor="#d9e2ec",
            lightcolor="#d90429",
            darkcolor="#d90429",
        )

        self.build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    # -------------------------
    # UI
    # -------------------------
    def build_ui(self):
        main = tk.Frame(self.root, bg="#edf2f7", padx=16, pady=16)
        main.pack(fill="both", expand=True)

        header = tk.Frame(main, bg="#0f172a")
        header.pack(fill="x", pady=(0, 12))
        header_inner = tk.Frame(header, bg="#0f172a", padx=20, pady=16)
        header_inner.pack(fill="x")

        tk.Label(
            header_inner,
            text="KSeF – Zestawienie FV do Excel",
            font=("Segoe UI", 24, "bold"),
            bg="#0f172a",
            fg="#ffffff",
        ).pack(anchor="w")

        tk.Label(
            header_inner,
            text="Logowanie ręczne do KSeF, wybór dat i eksport całej listy do tabeli Excel.",
            font=("Segoe UI", 10),
            bg="#0f172a",
            fg="#cbd5e1",
        ).pack(anchor="w", pady=(6, 0))

        stats_row = tk.Frame(main, bg="#edf2f7")
        stats_row.pack(fill="x", pady=(0, 10))
        self.make_stat_card(stats_row, "Wiersze", self.rows_var, "#0f4c81").pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.make_stat_card(stats_row, "Strony", self.pages_var, "#166534").pack(side="left", fill="x", expand=True, padx=8)
        self.make_stat_card(stats_row, "Ostatni plik", self.file_var, "#b45309").pack(side="left", fill="x", expand=True, padx=(8, 0))

        body = tk.Frame(main, bg="#edf2f7")
        body.pack(fill="both", expand=True)

        left = self.card(body, width=430)
        left.pack(side="left", fill="y", padx=(0, 10))
        left.pack_propagate(False)

        center = self.card(body)
        center.pack(side="left", fill="both", expand=True)

        tk.Label(left, text="Sterowanie", font=("Segoe UI", 16, "bold"), bg="white", fg="#0f172a").pack(anchor="w", pady=(0, 12))

        grid = tk.Frame(left, bg="white")
        grid.pack(fill="x", pady=(0, 12))
        grid.grid_columnconfigure(0, weight=1)
        grid.grid_columnconfigure(1, weight=1)

        ttk.Button(grid, text="Start / Otwórz KSeF", style="Primary.TButton", command=self.start_browser).grid(
            row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8)
        )
        ttk.Button(grid, text="Ustaw daty w KSeF", style="Secondary.TButton", command=self.apply_dates_in_ksef).grid(
            row=1, column=0, columnspan=2, sticky="ew", pady=(0, 8)
        )
        ttk.Button(grid, text="Sprawdź ilość FV", style="Secondary.TButton", command=self.check_invoice_count).grid(
            row=2, column=0, sticky="ew", padx=(0, 6), pady=(0, 8)
        )
        ttk.Button(grid, text="Eksport do Excel", style="Danger.TButton", command=self.export_summary).grid(
            row=2, column=1, sticky="ew", padx=(6, 0), pady=(0, 8)
        )
        ttk.Button(grid, text="Otwórz folder", style="Secondary.TButton", command=self.open_output_folder).grid(
            row=3, column=0, columnspan=2, sticky="ew"
        )

        tk.Frame(left, bg="#e2e8f0", height=1).pack(fill="x", pady=10)

        tk.Label(left, text="Zakres dat", font=("Segoe UI", 14, "bold"), bg="white", fg="#0f172a").pack(anchor="w", pady=(0, 10))
        tk.Label(left, text="Data od (YYYY-MM-DD)", font=("Segoe UI", 10), bg="white", fg="#334155").pack(anchor="w")
        tk.Entry(left, textvariable=self.date_from_var, font=("Segoe UI", 12, "bold"), bd=1, relief="solid").pack(fill="x", pady=(6, 10), ipady=4)
        tk.Label(left, text="Data do (YYYY-MM-DD)", font=("Segoe UI", 10), bg="white", fg="#334155").pack(anchor="w")
        tk.Entry(left, textvariable=self.date_to_var, font=("Segoe UI", 12, "bold"), bd=1, relief="solid").pack(fill="x", pady=(6, 10), ipady=4)

        tk.Frame(left, bg="#e2e8f0", height=1).pack(fill="x", pady=14)

        tk.Label(left, text="Postęp", font=("Segoe UI", 12, "bold"), bg="white", fg="#0f172a").pack(anchor="w", pady=(0, 8))
        self.progress = ttk.Progressbar(left, mode="determinate", maximum=100, style="Modern.Horizontal.TProgressbar")
        self.progress.pack(fill="x")

        tk.Label(left, textvariable=self.status_var, font=("Segoe UI", 10, "bold"), bg="white", fg="#334155", wraplength=370, justify="left").pack(anchor="w", pady=(10, 0))
        tk.Label(left, textvariable=self.result_var, font=("Segoe UI", 10, "bold"), bg="white", fg="#0f4c81", wraplength=370, justify="left").pack(anchor="w", pady=(10, 0))

        tk.Label(left, text="Uwagi", font=("Segoe UI", 12, "bold"), bg="white", fg="#0f172a").pack(anchor="w", pady=(18, 8))
        notes = (
            "1. Kliknij Start i zaloguj się ręcznie do KSeF.\n"
            "2. Przejdź do listy faktur zakupu.\n"
            "3. Możesz spróbować użyć przycisku 'Ustaw daty w KSeF'.\n"
            "4. Gdyby KSeF miał inny układ pól, ustaw filtry ręcznie na stronie.\n"
            "5. Eksport zapisze plik Excel oraz dodatkowo surowe dane z tabeli."
        )
        tk.Label(left, text=notes, font=("Segoe UI", 10), bg="white", fg="#475569", justify="left", wraplength=370).pack(anchor="w")

        tk.Label(center, text="Log operacji", font=("Segoe UI", 16, "bold"), bg="white", fg="#0f172a").pack(anchor="w", pady=(0, 10))
        log_frame = tk.Frame(center, bg="#0b1220", bd=0)
        log_frame.pack(fill="both", expand=True)

        self.status_box = tk.Text(
            log_frame,
            height=28,
            font=("Consolas", 10),
            bg="#0b1220",
            fg="#d7e3f4",
            insertbackground="white",
            bd=0,
            relief="flat",
            wrap="word",
            padx=12,
            pady=12,
        )
        self.status_box.pack(fill="both", expand=True)
        self.log("[INFO] Aplikacja uruchomiona.")
        self.log("[INFO] To jest wersja oparta na starym downloaderze FV, ale przerobiona pod zestawienie do Excel.")

    def card(self, parent, width=None):
        frame = tk.Frame(parent, bg="white", bd=1, relief="solid", padx=16, pady=16)
        if width:
            frame.configure(width=width)
        return frame

    def make_stat_card(self, parent, label, value_var, accent):
        card = tk.Frame(parent, bg="white", bd=1, relief="solid", padx=12, pady=10)
        strip = tk.Frame(card, bg=accent, height=4)
        strip.pack(fill="x", pady=(0, 10))
        tk.Label(card, text=label, font=("Segoe UI", 10, "bold"), bg="white", fg="#475569").pack(anchor="w")
        tk.Label(card, textvariable=value_var, font=("Segoe UI", 15, "bold"), bg="white", fg="#0f172a").pack(anchor="w", pady=(4, 0))
        return card

    # -------------------------
    # Helpers
    # -------------------------
    def log(self, text: str):
        self.status_box.configure(state="normal")
        self.status_box.insert("end", f"{text}\n")
        self.status_box.see("end")
        self.status_box.configure(state="disabled")
        self.root.update_idletasks()

    def set_status(self, text: str):
        self.status_var.set(f"Status: {text}")
        self.root.update_idletasks()

    def update_progress(self, current: int, total: int, text: str):
        total = max(1, total)
        current = max(0, min(current, total))
        self.progress.configure(maximum=total)
        self.progress["value"] = current
        percent = int((current / total) * 100)
        self.status_var.set(f"Status: {text} ({current}/{total}, {percent}%)")
        self.root.update_idletasks()

    def reset_progress(self, text: str):
        self.progress.configure(maximum=100)
        self.progress["value"] = 0
        self.set_status(text)

    def open_output_folder(self):
        try:
            os.makedirs(self.output_dir, exist_ok=True)
            os.startfile(self.output_dir)
        except Exception as e:
            messagebox.showerror("Błąd", f"Nie udało się otworzyć folderu.\n\n{e}")

    def normalize_spaces(self, text: str) -> str:
        return re.sub(r"\s+", " ", (text or "")).strip()

    def normalize_key(self, text: str) -> str:
        return re.sub(r"\s+", " ", (text or "")).strip().lower()

    def parse_date(self, value: str) -> str:
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        raise ValueError(f"Niepoprawna data: {value}")

    def safe_click_first(self, selectors: List[str], timeout: int = 5000, wait_after: int = 800) -> bool:
        for selector in selectors:
            try:
                loc = self.page.locator(selector)
                if loc.count() > 0 and loc.first.is_visible():
                    loc.first.click(timeout=timeout)
                    self.page.wait_for_timeout(wait_after)
                    self.log(f"[OK] Kliknięto: {selector}")
                    return True
            except Exception:
                pass
        return False

    def force_fill_first(self, selectors: List[str], value: str) -> bool:
        for selector in selectors:
            try:
                loc = self.page.locator(selector)
                if loc.count() == 0:
                    continue
                el = loc.first
                if not el.is_visible():
                    continue
                el.scroll_into_view_if_needed(timeout=3000)
                try:
                    el.click(timeout=2000)
                except Exception:
                    pass
                try:
                    el.fill(value, timeout=3000)
                except Exception:
                    try:
                        el.evaluate(
                            """
                            (node, val) => {
                                node.value = val;
                                node.dispatchEvent(new Event('input', { bubbles: true }));
                                node.dispatchEvent(new Event('change', { bubbles: true }));
                                node.dispatchEvent(new Event('blur', { bubbles: true }));
                            }
                            """,
                            value,
                        )
                    except Exception:
                        continue
                self.log(f"[OK] Ustawiono pole: {selector} = {value}")
                return True
            except Exception:
                pass
        return False

    def extract_item_key(self, text: str) -> str:
        text = self.normalize_spaces(text)
        patterns = [
            r"(\d{10,}-\d{8}-[A-Z0-9]+-\w+)",
            r"([A-Z0-9/\-]{6,}/\d{4})",
        ]
        for pattern in patterns:
            m = re.search(pattern, text, re.IGNORECASE)
            if m:
                return m.group(1).lower()
        return text.lower()

    def guess_cell_key(self, cells: List[str]) -> str:
        joined = " | ".join(cells)
        return self.extract_item_key(joined)

    def canonical_header(self, header: str) -> str:
        key = self.normalize_key(header)
        for target, aliases in HEADER_ALIASES.items():
            if key == self.normalize_key(target):
                return target
            if key in [self.normalize_key(a) for a in aliases]:
                return target
        return header.strip()

    def get_table_headers(self) -> List[str]:
        selectors = ["thead th", "[role='columnheader']"]
        for selector in selectors:
            try:
                loc = self.page.locator(selector)
                headers = []
                for i in range(loc.count()):
                    txt = self.normalize_spaces(loc.nth(i).inner_text())
                    if txt:
                        headers.append(txt)
                if headers:
                    return headers
            except Exception:
                pass
        return []

    def get_current_page_rows(self) -> List[Dict]:
        rows_data = []
        row_selectors = ["tbody tr", "table tbody tr", "[role='row']"]
        rows = None
        for selector in row_selectors:
            try:
                loc = self.page.locator(selector)
                if loc.count() > 0:
                    rows = loc
                    break
            except Exception:
                pass

        if rows is None:
            return rows_data

        count = rows.count()
        for i in range(count):
            try:
                row = rows.nth(i)
                if not row.is_visible():
                    continue

                cell_loc = row.locator("td, [role='cell']")
                cells = []
                for j in range(cell_loc.count()):
                    txt = self.normalize_spaces(cell_loc.nth(j).inner_text())
                    cells.append(txt)

                cells = [c for c in cells if c != ""]
                if len(cells) < 2:
                    continue

                row_id = self.guess_cell_key(cells)
                rows_data.append({
                    "cells": cells,
                    "row_id": row_id,
                    "row_text": " | ".join(cells),
                    "index": i,
                })
            except Exception:
                pass
        return rows_data

    def get_page_signature(self) -> str:
        rows = self.get_current_page_rows()
        if not rows:
            return "EMPTY"
        return "|".join(item["row_id"] for item in rows[:5])

    def go_to_next_page(self) -> bool:
        selectors = [
            "button[aria-label*='Następna']",
            "button[title*='Następna']",
            "[role='button'][aria-label*='Następna']",
            "text=Następna",
            "text=Next",
            "button:has-text('>')",
        ]
        for selector in selectors:
            try:
                loc = self.page.locator(selector)
                if loc.count() == 0:
                    continue
                btn = loc.first
                disabled = btn.get_attribute("disabled")
                aria_disabled = btn.get_attribute("aria-disabled")
                cls = (btn.get_attribute("class") or "").lower()
                if disabled is not None or aria_disabled == "true" or "disabled" in cls:
                    continue
                before = self.get_page_signature()
                btn.click(timeout=5000)
                self.page.wait_for_timeout(1800)
                after = self.get_page_signature()
                if after != before:
                    self.log("[OK] Przejście na następną stronę")
                    return True
            except Exception:
                pass
        return False

    def go_to_first_page(self, max_steps: int = 50):
        selectors = [
            "button[aria-label*='Poprzednia']",
            "button[title*='Poprzednia']",
            "[role='button'][aria-label*='Poprzednia']",
            "text=Poprzednia",
            "text=Previous",
            "button:has-text('<')",
        ]
        for _ in range(max_steps):
            moved = False
            for selector in selectors:
                try:
                    loc = self.page.locator(selector)
                    if loc.count() == 0:
                        continue
                    btn = loc.first
                    disabled = btn.get_attribute("disabled")
                    aria_disabled = btn.get_attribute("aria-disabled")
                    cls = (btn.get_attribute("class") or "").lower()
                    if disabled is not None or aria_disabled == "true" or "disabled" in cls:
                        continue
                    before = self.get_page_signature()
                    btn.click(timeout=3000)
                    self.page.wait_for_timeout(1200)
                    after = self.get_page_signature()
                    if after != before:
                        moved = True
                        break
                except Exception:
                    pass
            if not moved:
                break

    def scan_all_pages(self) -> Tuple[List[str], List[Dict], int]:
        if self.page is None:
            raise RuntimeError("Najpierw otwórz KSeF.")

        self.go_to_first_page()
        self.page.wait_for_timeout(1200)

        headers = self.get_table_headers()
        all_rows: List[Dict] = []
        seen_page_signatures = set()
        seen_rows = set()
        pages = 0

        while True:
            current_rows = self.get_current_page_rows()
            signature = self.get_page_signature()

            if signature in seen_page_signatures:
                break
            seen_page_signatures.add(signature)

            pages += 1
            self.log(f"[INFO] Odczytano stronę {pages}: {len(current_rows)} wierszy")

            for row in current_rows:
                if row["row_id"] in seen_rows:
                    continue
                seen_rows.add(row["row_id"])
                all_rows.append(row)

            self.update_progress(pages, max(1, pages + 1), "Skanowanie stron")

            if not self.go_to_next_page():
                break

        self.go_to_first_page()
        self.page.wait_for_timeout(1200)
        return headers, all_rows, pages

    def build_raw_headers(self, headers: List[str], rows: List[Dict]) -> List[str]:
        if headers and all(len(r["cells"]) == len(headers) for r in rows[: min(20, len(rows))] if r["cells"]):
            return headers
        max_len = max((len(r["cells"]) for r in rows), default=0)
        return [f"Kolumna {i + 1}" for i in range(max_len)]

    def map_to_target_rows(self, headers: List[str], rows: List[Dict]) -> List[Dict[str, str]]:
        raw_headers = self.build_raw_headers(headers, rows)
        canonical_headers = [self.canonical_header(h) for h in raw_headers]

        mapped = []
        for row in rows:
            cells = row["cells"]
            row_dict = {}
            for idx, header in enumerate(canonical_headers):
                row_dict[header] = cells[idx] if idx < len(cells) else ""

            normalized = {col: row_dict.get(col, "") for col in TARGET_COLUMNS}

            # Fallback: gdy kolumny nie są poprawnie rozpoznane, ale układ jest 11-elementowy.
            if not normalized["Nr KSeF"] and len(cells) >= 11:
                for idx, target in enumerate(TARGET_COLUMNS):
                    normalized[target] = cells[idx]

            mapped.append(normalized)
        return mapped

    def to_number_if_possible(self, value: str):
        if value is None:
            return ""
        if isinstance(value, (int, float)):
            return value
        text = str(value).strip().replace(" ", "")
        if not text:
            return ""
        text = text.replace(",", ".")
        if re.fullmatch(r"-?\d+(\.\d+)?", text):
            try:
                if "." in text:
                    return float(text)
                return int(text)
            except ValueError:
                return value
        return value

    def auto_fit_worksheet(self, ws):
        for col_cells in ws.columns:
            length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                value = "" if cell.value is None else str(cell.value)
                length = max(length, len(value))
            ws.column_dimensions[col_letter].width = min(max(length + 2, 12), 40)

    def style_header_row(self, ws, row_no=1):
        fill = PatternFill("solid", fgColor="0F4C81")
        font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
        for cell in ws[row_no]:
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def save_excel(self, headers: List[str], rows: List[Dict]) -> str:
        if not rows:
            raise ValueError("Brak danych do zapisania.")

        from_date = self.parse_date(self.date_from_var.get())
        to_date = self.parse_date(self.date_to_var.get())
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"KSeF_zestawienie_{from_date}_{to_date}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()

        ws_info = wb.active
        ws_info.title = "Info"
        ws_info.append(["Parametr", "Wartość"])
        ws_info.append(["Data od", from_date])
        ws_info.append(["Data do", to_date])
        ws_info.append(["Data eksportu", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws_info.append(["Liczba wierszy", len(rows)])
        ws_info.append(["Liczba stron", self.last_pages])
        self.style_header_row(ws_info)
        self.auto_fit_worksheet(ws_info)

        ws_target = wb.create_sheet("Zestawienie")
        mapped_rows = self.map_to_target_rows(headers, rows)
        ws_target.append(TARGET_COLUMNS)
        for row in mapped_rows:
            ws_target.append([
                row["Identyfikator sprzedawcy"],
                row["Nazwa sprzedawcy"],
                row["Nr KSeF"],
                row["Nr faktury"],
                row["Data wystawienia"],
                row["Data zapisania w KSeF"],
                row["Data otrzymania"],
                row["Waluta"],
                self.to_number_if_possible(row["Netto"]),
                self.to_number_if_possible(row["Brutto"]),
                self.to_number_if_possible(row["VAT (PLN)"]),
            ])
        self.style_header_row(ws_target)
        self.auto_fit_worksheet(ws_target)
        ws_target.freeze_panes = "A2"

        raw_headers = self.build_raw_headers(headers, rows)
        ws_raw = wb.create_sheet("Surowe dane")
        ws_raw.append(raw_headers)
        for row in rows:
            line = row["cells"] + [""] * (len(raw_headers) - len(row["cells"]))
            ws_raw.append(line[: len(raw_headers)])
        self.style_header_row(ws_raw)
        self.auto_fit_worksheet(ws_raw)
        ws_raw.freeze_panes = "A2"

        wb.save(filepath)
        return filepath

    # -------------------------
    # Browser
    # -------------------------
    def start_browser(self):
        try:
            if self.page is not None:
                messagebox.showinfo("Informacja", "Przeglądarka jest już otwarta.")
                return

            self.reset_progress("Uruchamianie przeglądarki")
            self.log("[INFO] Uruchamianie przeglądarki...")

            self.playwright = sync_playwright().start()
            self.browser = self.playwright.chromium.launch(headless=False, slow_mo=100)
            self.context = self.browser.new_context(accept_downloads=True)
            self.page = self.context.new_page()
            self.page.goto("https://ap.ksef.mf.gov.pl/web/invoice-list", timeout=90000)

            self.reset_progress("Czekam na logowanie")
            self.log("[OK] KSeF otwarty.")
            self.log("[INFO] Zaloguj się ręcznie i przejdź do listy faktur zakupu.")
            messagebox.showinfo(
                "Logowanie",
                "KSeF został otwarty.\n\nZaloguj się ręcznie i przejdź do listy faktur zakupu.",
            )
        except Exception as e:
            self.reset_progress("Błąd")
            self.log(f"[BŁĄD] Nie udało się otworzyć przeglądarki: {e}")
            messagebox.showerror("Błąd", f"Nie udało się otworzyć przeglądarki.\n\n{e}")

    def apply_dates_in_ksef(self):
        try:
            if self.page is None:
                messagebox.showwarning("Uwaga", "Najpierw kliknij Start.")
                return

            date_from = self.parse_date(self.date_from_var.get())
            date_to = self.parse_date(self.date_to_var.get())

            self.reset_progress("Ustawianie dat w KSeF")
            self.log(f"[INFO] Próba ustawienia dat: {date_from} -> {date_to}")

            from_selectors = [
                "input[placeholder*='Od']",
                "input[placeholder*='od']",
                "input[aria-label*='Od']",
                "input[aria-label*='od']",
                "input[name*='from']",
                "input[name*='From']",
                "input[id*='from']",
                "input[id*='From']",
            ]
            to_selectors = [
                "input[placeholder*='Do']",
                "input[placeholder*='do']",
                "input[aria-label*='Do']",
                "input[aria-label*='do']",
                "input[name*='to']",
                "input[name*='To']",
                "input[id*='to']",
                "input[id*='To']",
            ]

            ok_from = self.force_fill_first(from_selectors, date_from)
            ok_to = self.force_fill_first(to_selectors, date_to)

            clicked = self.safe_click_first(
                [
                    "button:has-text('Zastosuj')",
                    "button:has-text('Filtruj')",
                    "button:has-text('Szukaj')",
                    "text=Zastosuj",
                    "text=Filtruj",
                ],
                timeout=3000,
                wait_after=1400,
            )

            if ok_from and ok_to:
                self.reset_progress("Daty ustawione")
                if clicked:
                    self.log("[OK] Daty ustawione i zatwierdzone.")
                else:
                    self.log("[OK] Daty ustawione. Jeśli lista się nie odświeżyła, kliknij filtr ręcznie w KSeF.")
                messagebox.showinfo(
                    "Gotowe",
                    "Daty zostały wpisane.\n\nJeśli lista nie odświeżyła się sama, kliknij filtr ręcznie w KSeF.",
                )
            else:
                self.reset_progress("Nie znaleziono pól dat")
                self.log("[INFO] Nie udało się automatycznie znaleźć pól dat w KSeF.")
                messagebox.showwarning(
                    "Uwaga",
                    "Nie udało się automatycznie ustawić pól dat.\n\nUstaw daty ręcznie na stronie KSeF i potem kliknij 'Eksport do Excel'.",
                )
        except Exception as e:
            self.reset_progress("Błąd ustawiania dat")
            self.log(f"[BŁĄD] Nie udało się ustawić dat: {e}")
            messagebox.showerror("Błąd", f"Nie udało się ustawić dat.\n\n{e}")

    # -------------------------
    # Scan / Export
    # -------------------------
    def check_invoice_count(self):
        try:
            if self.page is None:
                messagebox.showwarning("Uwaga", "Najpierw kliknij Start.")
                return

            self.reset_progress("Skanowanie stron")
            self.log("[INFO] Skanuję wszystkie strony z listą FV...")

            headers, rows, pages = self.scan_all_pages()
            self.last_headers = headers
            self.last_rows = rows
            self.last_pages = pages

            self.rows_var.set(str(len(rows)))
            self.pages_var.set(str(pages))
            self.result_var.set(f"Wynik: znaleziono {len(rows)} wierszy na {pages} stronach")
            self.reset_progress("Sprawdzanie zakończone")

            if headers:
                self.log(f"[INFO] Wykryte nagłówki: {' | '.join(headers)}")
            self.log(f"[OK] Wszystkich wierszy: {len(rows)}")
            self.log(f"[OK] Liczba stron: {pages}")

            messagebox.showinfo("Wynik", f"Znaleziono {len(rows)} wierszy na {pages} stronach.")
        except Exception as e:
            self.reset_progress("Błąd liczenia")
            self.log(f"[BŁĄD] Nie udało się sprawdzić ilości FV: {e}")
            messagebox.showerror("Błąd", f"Nie udało się sprawdzić ilości FV.\n\n{e}")

    def export_summary(self):
        try:
            if self.page is None:
                messagebox.showwarning("Uwaga", "Najpierw kliknij Start.")
                return

            if not self.last_rows:
                self.log("[INFO] Brak zapisanych wyników skanowania. Skanuję teraz listę...")
                headers, rows, pages = self.scan_all_pages()
                self.last_headers = headers
                self.last_rows = rows
                self.last_pages = pages
                self.rows_var.set(str(len(rows)))
                self.pages_var.set(str(pages))

            self.update_progress(1, 3, "Przygotowanie eksportu")
            filepath = self.save_excel(self.last_headers, self.last_rows)
            self.update_progress(3, 3, "Eksport zakończony")

            self.file_var.set(Path(filepath).name)
            self.result_var.set(f"Wynik: zapisano {len(self.last_rows)} wierszy do pliku Excel")
            self.log(f"[OK] Zapisano plik: {filepath}")

            messagebox.showinfo(
                "Sukces",
                f"Zestawienie zostało zapisane.\n\nPlik:\n{filepath}",
            )
        except Exception as e:
            self.reset_progress("Błąd eksportu")
            self.log(f"[BŁĄD] Nie udało się zapisać Excel: {e}")
            messagebox.showerror("Błąd", f"Nie udało się zapisać Excel.\n\n{e}")

    # -------------------------
    # Close
    # -------------------------
    def on_close(self):
        try:
            if self.context:
                self.context.close()
        except Exception:
            pass
        try:
            if self.browser:
                self.browser.close()
        except Exception:
            pass
        try:
            if self.playwright:
                self.playwright.stop()
        except Exception:
            pass
        self.root.destroy()


if __name__ == "__main__":
    if sys.platform.startswith("win"):
        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

    root = tk.Tk()
    app = KsefSummaryApp(root)
    root.mainloop()
