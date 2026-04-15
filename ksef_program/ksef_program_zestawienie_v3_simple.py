import os
import re
import sys
import math
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple

try:
    from PIL import Image, ImageTk
except Exception:
    Image = None
    ImageTk = None

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

TARGET_COLUMNS = [
    "Identyfikator sprzedawcy",
    "Nazwa sprzedawcy",
    "Nr KSeF",
    "Nr faktury",
    "Data wystawienia",
    "Data zapisania w KSeF",
    "Data otrzymania",
    "Waluta",
    "Netto",
    "Brutto",
    "VAT (PLN)",
]

HEADER_ALIASES = {
    "Identyfikator sprzedawcy": ["identyfikator sprzedawcy", "nip sprzedawcy", "nip", "sprzedawca nip"],
    "Nazwa sprzedawcy": ["nazwa sprzedawcy", "sprzedawca", "nazwa podmiotu", "nazwa"],
    "Nr KSeF": ["nr ksef", "numer ksef", "ksef", "numer referencyjny ksef"],
    "Nr faktury": ["nr faktury", "numer faktury", "faktura"],
    "Data wystawienia": ["data wystawienia", "wystawiono"],
    "Data zapisania w KSeF": ["data zapisania w ksef", "data przyjęcia w ksef", "zapisano w ksef"],
    "Data otrzymania": ["data otrzymania", "otrzymano"],
    "Waluta": ["waluta"],
    "Netto": ["netto", "wartość netto", "kwota netto"],
    "Brutto": ["brutto", "wartość brutto", "kwota brutto"],
    "VAT (PLN)": ["vat (pln)", "vat pln", "vat", "kwota vat pln"],
}


class KsefSimpleSummaryApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("KSeF Program do Arkuszy")
        self.root.geometry("980x700")
        self.root.minsize(900, 640)
        self.root.configure(bg="#f5f7fb")

        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None

        self.base_dir = str(Path(__file__).resolve().parent)
        self.output_dir = os.path.join(self.base_dir, "zestawienia_ksef")
        os.makedirs(self.output_dir, exist_ok=True)

        today = datetime.today()
        month_start = today.replace(day=1)

        self.date_from_var = tk.StringVar(value=month_start.strftime("%Y-%m-%d"))
        self.date_to_var = tk.StringVar(value=today.strftime("%Y-%m-%d"))
        self.status_var = tk.StringVar(value="Otwórz KSeF")
        self.summary_var = tk.StringVar(value="Brak zapisanych danych")
        self.file_var = tk.StringVar(value="-")

        self.last_headers: List[str] = []
        self.last_rows: List[Dict] = []
        self.last_pages: int = 0
        self.logo_image = None
        self.busy_running = False
        self.dot_phase = 0
        self.dot_colors_idle = ["#cbd5e1", "#94a3b8", "#64748b", "#94a3b8"]
        self.dot_colors_ready = ["#22c55e", "#4ade80", "#86efac", "#4ade80"]
        self.dot_colors_busy = ["#f97316", "#fb923c", "#fdba74", "#fb923c"]

        self.setup_style()
        self.build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)


    def setup_style(self):
        style = ttk.Style()
        style.theme_use("clam")

        style.configure(
            "Main.TButton",
            font=("Segoe UI", 11, "bold"),
            padding=(16, 11),
            foreground="#ffffff",
            background="#c81f25",
            borderwidth=0,
        )
        style.map(
            "Main.TButton",
            background=[("active", "#a9151b"), ("pressed", "#8f1217")],
            foreground=[("disabled", "#f3f4f6")],
        )

        style.configure(
            "Ghost.TButton",
            font=("Segoe UI", 10, "bold"),
            padding=(14, 10),
            foreground="#0f172a",
            background="#ffffff",
            borderwidth=1,
            relief="solid",
        )
        style.map(
            "Ghost.TButton",
            background=[("active", "#eef2f7"), ("pressed", "#e2e8f0")],
        )

        style.configure(
            "Red.Horizontal.TProgressbar",
            troughcolor="#e2e8f0",
            background="#c81f25",
            bordercolor="#e2e8f0",
            lightcolor="#c81f25",
            darkcolor="#c81f25",
            thickness=10,
        )

    def find_logo_file(self) -> str:
        graphics_dir = Path(self.base_dir) / "Grafiki"
        if not graphics_dir.exists():
            return ""

        preferred_names = [
            "logo.png", "logo.gif", "logo.jpg", "logo.jpeg", "logo.webp",
            "Logo.png", "Logo.gif", "Logo.jpg", "Logo.jpeg", "Logo.webp",
        ]
        for name in preferred_names:
            candidate = graphics_dir / name
            if candidate.exists():
                return str(candidate)

        for ext in ("*.png", "*.gif", "*.jpg", "*.jpeg", "*.webp"):
            matches = sorted(graphics_dir.glob(ext))
            if matches:
                return str(matches[0])

        return ""


    def load_logo(self):
        logo_path = self.find_logo_file()
        if not logo_path:
            return None

        max_w, max_h = 280, 78

        try:
            if Image is not None and ImageTk is not None:
                img = Image.open(logo_path)
                img.thumbnail((max_w, max_h))
                return ImageTk.PhotoImage(img)

            suffix = Path(logo_path).suffix.lower()
            if suffix in {".png", ".gif"}:
                img = tk.PhotoImage(file=logo_path)
                width = max(1, img.width())
                height = max(1, img.height())
                scale_w = math.ceil(width / max_w)
                scale_h = math.ceil(height / max_h)
                scale = max(1, scale_w, scale_h)
                if scale > 1:
                    img = img.subsample(scale, scale)
                return img
        except Exception:
            return None

        return None


    def build_ui(self):
        outer = tk.Frame(self.root, bg="#eef2f7", padx=18, pady=18)
        outer.pack(fill="both", expand=True)

        header = tk.Frame(outer, bg="#ffffff", bd=0, highlightbackground="#d8dee8", highlightthickness=1)
        header.pack(fill="x", pady=(0, 14))

        accent = tk.Frame(header, bg="#c81f25", height=5)
        accent.pack(fill="x", side="top")

        header_inner = tk.Frame(header, bg="#ffffff", padx=22, pady=18)
        header_inner.pack(fill="x")

        left_header = tk.Frame(header_inner, bg="#ffffff")
        left_header.pack(side="left", fill="both", expand=True)

        tk.Label(
            left_header,
            text="KSeF Program do Arkuszy",
            font=("Segoe UI", 22, "bold"),
            bg="#ffffff",
            fg="#0f172a",
        ).pack(anchor="w")

        tk.Label(
            left_header,
            text="by Paweł Ruchlicki",
            font=("Segoe UI", 10),
            bg="#ffffff",
            fg="#64748b",
        ).pack(anchor="w", pady=(6, 0))

        self.logo_image = self.load_logo()
        if self.logo_image is not None:
            logo_wrap = tk.Frame(header_inner, bg="#ffffff")
            logo_wrap.pack(side="right", anchor="ne", padx=(18, 0))

            logo_box = tk.Frame(logo_wrap, bg="#ffffff", bd=0, highlightbackground="#e2e8f0", highlightthickness=1, padx=14, pady=10)
            logo_box.pack(anchor="e")

            tk.Label(logo_box, image=self.logo_image, bg="#ffffff").pack(anchor="e")

        controls = tk.Frame(outer, bg="#eef2f7")
        controls.pack(fill="x", pady=(0, 14))

        left_panel = tk.Frame(controls, bg="#ffffff", bd=0, highlightbackground="#d8dee8", highlightthickness=1, padx=18, pady=16)
        left_panel.pack(side="left", fill="x", expand=True)

        tk.Label(left_panel, text="Zakres do nazwy pliku", font=("Segoe UI", 10, "bold"), bg="#ffffff", fg="#334155").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 10))

        tk.Label(left_panel, text="Data od", font=("Segoe UI", 10, "bold"), bg="#ffffff", fg="#0f172a").grid(row=1, column=0, sticky="w")
        date_from_entry = tk.Entry(left_panel, textvariable=self.date_from_var, font=("Segoe UI", 12), relief="solid", bd=1, width=16)
        date_from_entry.grid(row=2, column=0, sticky="w", padx=(0, 14), ipady=6, pady=(4, 0))

        tk.Label(left_panel, text="Data do", font=("Segoe UI", 10, "bold"), bg="#ffffff", fg="#0f172a").grid(row=1, column=1, sticky="w")
        date_to_entry = tk.Entry(left_panel, textvariable=self.date_to_var, font=("Segoe UI", 12), relief="solid", bd=1, width=16)
        date_to_entry.grid(row=2, column=1, sticky="w", ipady=6, pady=(4, 0))

        buttons_panel = tk.Frame(controls, bg="#ffffff", bd=0, highlightbackground="#d8dee8", highlightthickness=1, padx=18, pady=16)
        buttons_panel.pack(side="left", padx=(12, 0))

        ttk.Button(buttons_panel, text="Otwórz KSeF", style="Ghost.TButton", command=self.start_browser).pack(fill="x")
        ttk.Button(buttons_panel, text="Pobierz zestawienie", style="Main.TButton", command=self.run_full_export).pack(fill="x", pady=10)
        ttk.Button(buttons_panel, text="Otwórz folder", style="Ghost.TButton", command=self.open_output_folder).pack(fill="x")

        status_row = tk.Frame(outer, bg="#eef2f7")
        status_row.pack(fill="x", pady=(0, 14))

        self.status_card = tk.Frame(status_row, bg="#ffffff", bd=0, highlightbackground="#d8dee8", highlightthickness=1, padx=16, pady=14)
        self.status_card.pack(side="left", fill="x", expand=True)

        self.result_card = tk.Frame(status_row, bg="#ffffff", bd=0, highlightbackground="#d8dee8", highlightthickness=1, padx=16, pady=14)
        self.result_card.pack(side="left", fill="x", expand=True, padx=12)

        self.file_card = tk.Frame(status_row, bg="#ffffff", bd=0, highlightbackground="#d8dee8", highlightthickness=1, padx=16, pady=14)
        self.file_card.pack(side="left", fill="x", expand=True)

        tk.Label(self.status_card, text="Status", font=("Segoe UI", 10, "bold"), bg="#ffffff", fg="#334155").pack(anchor="w")
        status_line = tk.Frame(self.status_card, bg="#ffffff")
        status_line.pack(anchor="w", fill="x", pady=(10, 0))

        self.status_dot_canvas = tk.Canvas(status_line, width=14, height=14, bg="#ffffff", highlightthickness=0, bd=0)
        self.status_dot_canvas.pack(side="left")
        self.status_dot_item = self.status_dot_canvas.create_oval(2, 2, 12, 12, fill="#94a3b8", outline="")

        self.status_label = tk.Label(status_line, textvariable=self.status_var, font=("Segoe UI", 12, "bold"), bg="#ffffff", fg="#0f172a")
        self.status_label.pack(side="left", padx=(8, 0))

        tk.Label(self.result_card, text="Wynik", font=("Segoe UI", 10, "bold"), bg="#ffffff", fg="#334155").pack(anchor="w")
        tk.Label(self.result_card, textvariable=self.summary_var, font=("Segoe UI", 12), bg="#ffffff", fg="#0f172a", wraplength=250, justify="left").pack(anchor="w", pady=(10, 0))

        tk.Label(self.file_card, text="Plik", font=("Segoe UI", 10, "bold"), bg="#ffffff", fg="#334155").pack(anchor="w")
        tk.Label(self.file_card, textvariable=self.file_var, font=("Segoe UI", 12), bg="#ffffff", fg="#0f172a", wraplength=250, justify="left").pack(anchor="w", pady=(10, 0))

        progress_wrap = tk.Frame(outer, bg="#eef2f7")
        progress_wrap.pack(fill="x", pady=(0, 14))

        self.progress = ttk.Progressbar(progress_wrap, mode="determinate", maximum=100, style="Red.Horizontal.TProgressbar")
        self.progress.pack(fill="x")

        log_card = tk.Frame(outer, bg="#ffffff", bd=0, highlightbackground="#d8dee8", highlightthickness=1)
        log_card.pack(fill="both", expand=True)

        log_header = tk.Frame(log_card, bg="#ffffff")
        log_header.pack(fill="x", padx=16, pady=(14, 8))

        tk.Label(log_header, text="Log", font=("Segoe UI", 14, "bold"), bg="#ffffff", fg="#0f172a").pack(side="left")

        self.log_box = tk.Text(
            log_card,
            font=("Consolas", 10),
            bg="#081229",
            fg="#e2e8f0",
            insertbackground="#ffffff",
            wrap="word",
            bd=0,
            relief="flat",
            padx=14,
            pady=14,
            height=18,
        )
        self.log_box.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        self.log("[INFO] Aplikacja uruchomiona.")
        self.log("[INFO] Ustaw filtry ręcznie w KSeF, a potem kliknij pobierz zestawienie.")
        self.animate_status_dot()

    def log(self, text: str):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", text + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")
        self.root.update_idletasks()

    def set_status(self, text: str):
        self.status_var.set(text)
        self.root.update_idletasks()

    def update_progress(self, current: int, total: int, text: str):
        total = max(1, total)
        current = max(0, min(current, total))
        self.busy_running = False
        self.progress.stop()
        self.progress.configure(mode="determinate", maximum=total)
        self.progress["value"] = current
        percent = int(current / total * 100)
        self.set_status(f"{text} ({percent}%)")

    def reset_progress(self, text: str):
        self.busy_running = False
        self.progress.stop()
        self.progress.configure(mode="determinate", maximum=100)
        self.progress["value"] = 0
        self.set_status(text)


    def start_busy(self, text: str):
        self.busy_running = True
        self.progress.stop()
        self.progress.configure(mode="indeterminate")
        self.progress.start(12)
        self.set_status(text)

    def stop_busy(self, text: str, progress_value: int = 0):
        self.busy_running = False
        self.progress.stop()
        self.progress.configure(mode="determinate", maximum=100)
        self.progress["value"] = progress_value
        self.set_status(text)

    def animate_status_dot(self):
        if hasattr(self, "status_dot_canvas"):
            if self.busy_running:
                palette = self.dot_colors_busy
            elif self.page is not None:
                palette = self.dot_colors_ready
            else:
                palette = self.dot_colors_idle

            color = palette[self.dot_phase % len(palette)]
            self.status_dot_canvas.itemconfigure(self.status_dot_item, fill=color)
            self.dot_phase += 1

        self.root.after(180, self.animate_status_dot)

    def open_output_folder(self):
        try:
            os.makedirs(self.output_dir, exist_ok=True)
            os.startfile(self.output_dir)
        except Exception as e:
            messagebox.showerror("Błąd", f"Nie udało się otworzyć folderu.\n\n{e}")

    def close_browser_resources(self):
        try:
            if self.context:
                self.context.close()
        except Exception:
            pass
        try:
            if self.browser:
                self.browser.close()
        except Exception:
            pass
        try:
            if self.playwright:
                self.playwright.stop()
        except Exception:
            pass

        self.context = None
        self.browser = None
        self.page = None
        self.playwright = None

    def parse_date(self, value: str) -> str:
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        raise ValueError(f"Niepoprawna data: {value}")

    def normalize_spaces(self, text: str) -> str:
        return re.sub(r"\s+", " ", (text or "")).strip()

    def normalize_key(self, text: str) -> str:
        return self.normalize_spaces(text).lower()

    def canonical_header(self, header: str) -> str:
        key = self.normalize_key(header)
        for target, aliases in HEADER_ALIASES.items():
            if key == self.normalize_key(target):
                return target
            if key in [self.normalize_key(a) for a in aliases]:
                return target
        return header.strip()

    def validate_date_range(self) -> Tuple[str, str]:
        date_from = self.parse_date(self.date_from_var.get())
        date_to = self.parse_date(self.date_to_var.get())
        if date_from > date_to:
            raise ValueError("Data od nie może być późniejsza niż data do.")
        return date_from, date_to

    def is_browser_ready(self) -> bool:
        if self.page is None:
            return False
        try:
            return not self.page.is_closed()
        except Exception:
            return False

    def ensure_browser_ready(self):
        if self.is_browser_ready():
            return
        self.close_browser_resources()
        raise RuntimeError("Przeglądarka KSeF nie jest otwarta. Kliknij 'Otwórz KSeF' i zaloguj się ponownie.")

    def wait_for_table_ready(self, timeout_ms: int = 15000) -> bool:
        selectors = [
            "tbody tr",
            "table tbody tr",
            "[role='row']",
            "thead th",
            "[role='columnheader']",
        ]
        for selector in selectors:
            try:
                self.page.wait_for_selector(selector, timeout=timeout_ms)
                return True
            except Exception:
                pass
        return False

    def start_browser(self):
        try:
            if self.is_browser_ready():
                messagebox.showinfo("Informacja", "Przeglądarka jest już otwarta.")
                return

            self.close_browser_resources()

            self.start_busy("Uruchamianie przeglądarki")
            self.log("[INFO] Uruchamianie przeglądarki...")

            self.playwright = sync_playwright().start()
            self.browser = self.playwright.chromium.launch(headless=False, slow_mo=100)
            self.context = self.browser.new_context(accept_downloads=True)
            self.page = self.context.new_page()
            self.page.goto("https://ap.ksef.mf.gov.pl/web/invoice-list", timeout=90000)

            self.stop_busy("Zaloguj się do KSeF")
            self.log("[OK] KSeF otwarty.")
            self.log("[INFO] Zaloguj się ręcznie, rozwiń filtry i ustaw daty w KSeF.")
            self.log("[INFO] Potem przejdź do listy faktur zakupu i kliknij pobierz zestawienie.")
            messagebox.showinfo(
                "KSeF",
                "KSeF został otwarty.\n\nZaloguj się ręcznie, rozwiń filtry, ustaw daty i przejdź do listy faktur zakupu.",
            )
        except Exception as e:
            self.close_browser_resources()
            self.reset_progress("Błąd")
            self.log(f"[BŁĄD] Nie udało się otworzyć przeglądarki: {e}")
            messagebox.showerror("Błąd", f"Nie udało się otworzyć przeglądarki.\n\n{e}")

    def force_fill_first(self, selectors: List[str], value: str) -> bool:
        for selector in selectors:
            try:
                loc = self.page.locator(selector)
                if loc.count() == 0:
                    continue
                el = loc.first
                if not el.is_visible():
                    continue
                el.scroll_into_view_if_needed(timeout=3000)
                try:
                    el.click(timeout=2000)
                except Exception:
                    pass
                try:
                    el.fill(value, timeout=3000)
                except Exception:
                    try:
                        el.evaluate(
                            """
                            (node, val) => {
                                node.value = val;
                                node.dispatchEvent(new Event('input', { bubbles: true }));
                                node.dispatchEvent(new Event('change', { bubbles: true }));
                                node.dispatchEvent(new Event('blur', { bubbles: true }));
                            }
                            """,
                            value,
                        )
                    except Exception:
                        continue
                return True
            except Exception:
                pass
        return False

    def safe_click_first(self, selectors: List[str], timeout: int = 4000, wait_after: int = 1200) -> bool:
        for selector in selectors:
            try:
                loc = self.page.locator(selector)
                if loc.count() > 0 and loc.first.is_visible():
                    loc.first.click(timeout=timeout)
                    self.page.wait_for_timeout(wait_after)
                    return True
            except Exception:
                pass
        return False

    def try_apply_dates(self):
        date_from, date_to = self.validate_date_range()

        from_selectors = [
            "input[placeholder*='Od']", "input[placeholder*='od']",
            "input[aria-label*='Od']", "input[aria-label*='od']",
            "input[name*='from']", "input[id*='from']",
        ]
        to_selectors = [
            "input[placeholder*='Do']", "input[placeholder*='do']",
            "input[aria-label*='Do']", "input[aria-label*='do']",
            "input[name*='to']", "input[id*='to']",
        ]

        ok_from = self.force_fill_first(from_selectors, date_from)
        ok_to = self.force_fill_first(to_selectors, date_to)
        clicked = self.safe_click_first([
            "button:has-text('Zastosuj')",
            "button:has-text('Filtruj')",
            "button:has-text('Szukaj')",
            "text=Zastosuj",
            "text=Filtruj",
        ])

        if ok_from and ok_to:
            if clicked:
                self.log(f"[OK] Ustawiono daty {date_from} - {date_to} i odświeżono listę.")
            else:
                self.log(f"[OK] Wpisano daty {date_from} - {date_to}. Jeśli lista się nie odświeżyła, kliknij filtr ręcznie.")
        else:
            self.log("[INFO] Nie udało się automatycznie znaleźć pól dat. Program pobierze to, co aktualnie widać na liście.")

        self.wait_for_table_ready(timeout_ms=10000)

    def get_table_headers(self) -> List[str]:
        for selector in ["thead th", "[role='columnheader']"]:
            try:
                loc = self.page.locator(selector)
                headers = []
                for i in range(loc.count()):
                    txt = self.normalize_spaces(loc.nth(i).inner_text())
                    if txt:
                        headers.append(txt)
                if headers:
                    return headers
            except Exception:
                pass
        return []

    def extract_item_key(self, text: str) -> str:
        text = self.normalize_spaces(text)
        patterns = [r"(\d{10,}-\d{8}-[A-Z0-9]+-\w+)", r"([A-Z0-9/\-]{6,}/\d{4})"]
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1).lower()
        return text.lower()

    def get_current_page_rows(self) -> List[Dict]:
        rows_data = []
        rows = None
        for selector in ["tbody tr", "table tbody tr", "[role='row']"]:
            try:
                loc = self.page.locator(selector)
                if loc.count() > 0:
                    rows = loc
                    break
            except Exception:
                pass

        if rows is None:
            return rows_data

        for i in range(rows.count()):
            try:
                row = rows.nth(i)
                if not row.is_visible():
                    continue
                cell_loc = row.locator("td, [role='cell']")
                cells = []
                for j in range(cell_loc.count()):
                    txt = self.normalize_spaces(cell_loc.nth(j).inner_text())
                    if txt:
                        cells.append(txt)
                if len(cells) < 2:
                    continue
                rows_data.append({
                    "cells": cells,
                    "row_id": self.extract_item_key(" | ".join(cells)),
                })
            except Exception:
                pass
        return rows_data

    def get_page_signature(self) -> str:
        rows = self.get_current_page_rows()
        if not rows:
            return "EMPTY"
        return "|".join(row["row_id"] for row in rows[:5])

    def go_to_first_page(self, max_steps: int = 50):
        selectors = [
            "button[aria-label*='Poprzednia']",
            "button[title*='Poprzednia']",
            "[role='button'][aria-label*='Poprzednia']",
            "text=Poprzednia",
            "text=Previous",
            "button:has-text('<')",
        ]
        for _ in range(max_steps):
            moved = False
            for selector in selectors:
                try:
                    loc = self.page.locator(selector)
                    if loc.count() == 0:
                        continue
                    btn = loc.first
                    disabled = btn.get_attribute("disabled")
                    aria_disabled = btn.get_attribute("aria-disabled")
                    classes = (btn.get_attribute("class") or "").lower()
                    if disabled is not None or aria_disabled == "true" or "disabled" in classes:
                        continue
                    before = self.get_page_signature()
                    btn.click(timeout=3000)
                    self.page.wait_for_timeout(1200)
                    after = self.get_page_signature()
                    if after != before:
                        moved = True
                        break
                except Exception:
                    pass
            if not moved:
                break

    def go_to_next_page(self) -> bool:
        selectors = [
            "button[aria-label*='Następna']",
            "button[title*='Następna']",
            "[role='button'][aria-label*='Następna']",
            "text=Następna",
            "text=Next",
            "button:has-text('>')",
        ]
        for selector in selectors:
            try:
                loc = self.page.locator(selector)
                if loc.count() == 0:
                    continue
                btn = loc.first
                disabled = btn.get_attribute("disabled")
                aria_disabled = btn.get_attribute("aria-disabled")
                classes = (btn.get_attribute("class") or "").lower()
                if disabled is not None or aria_disabled == "true" or "disabled" in classes:
                    continue
                before = self.get_page_signature()
                btn.click(timeout=5000)
                self.page.wait_for_timeout(1800)
                after = self.get_page_signature()
                if after != before:
                    return True
            except Exception:
                pass
        return False

    def scan_all_pages(self) -> Tuple[List[str], List[Dict], int]:
        self.go_to_first_page()
        self.page.wait_for_timeout(1200)

        headers = self.get_table_headers()
        all_rows: List[Dict] = []
        seen_pages = set()
        seen_rows = set()
        pages = 0

        while True:
            current_rows = self.get_current_page_rows()
            signature = self.get_page_signature()
            if signature in seen_pages:
                break
            seen_pages.add(signature)
            pages += 1
            self.log(f"[INFO] Odczyt strony {pages}: {len(current_rows)} wierszy")

            for row in current_rows:
                if row["row_id"] in seen_rows:
                    continue
                seen_rows.add(row["row_id"])
                all_rows.append(row)

            self.update_progress(pages, max(1, pages + 1), "Skanowanie listy")
            if not self.go_to_next_page():
                break

        self.go_to_first_page()
        self.page.wait_for_timeout(1200)
        return headers, all_rows, pages

    def build_raw_headers(self, headers: List[str], rows: List[Dict]) -> List[str]:
        if headers and all(len(r["cells"]) == len(headers) for r in rows[: min(20, len(rows))] if r["cells"]):
            return headers
        max_len = max((len(r["cells"]) for r in rows), default=0)
        return [f"Kolumna {i + 1}" for i in range(max_len)]

    def clean_ksef_number(self, value: str) -> str:
        text = self.normalize_spaces(value)
        match = re.search(r"(\d{10,}-\d{8}-[A-Z0-9]+-[A-Z0-9]+)", text, re.IGNORECASE)
        return match.group(1) if match else text

    def clean_invoice_number(self, value: str) -> str:
        text = self.normalize_spaces(value)
        text = re.sub(r"\s*Kopiuj numer faktury.*$", "", text, flags=re.IGNORECASE).strip()
        text = re.sub(r"\s*content_copy.*$", "", text, flags=re.IGNORECASE).strip()
        return text

    def normalize_export_date(self, value: str) -> str:
        text = self.normalize_spaces(value)
        for fmt_in in ("%d.%m.%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt_in).strftime("%Y-%m-%d")
            except ValueError:
                pass
        return text

    def parse_amount_and_currency(self, value: str) -> Tuple[object, str]:
        text = self.normalize_spaces(value)
        match = re.search(r"(-?[0-9\s.,]+)\s*([A-Z]{3})\b", text)
        if not match:
            return self.to_number_if_possible(text), ""
        amount_text = match.group(1).replace(" ", "")
        if "," in amount_text and "." in amount_text:
            amount_text = amount_text.replace(".", "").replace(",", ".")
        elif "," in amount_text:
            amount_text = amount_text.replace(",", ".")
        try:
            amount = float(amount_text)
            if amount.is_integer():
                amount = int(amount)
        except ValueError:
            amount = self.to_number_if_possible(match.group(1))
        return amount, match.group(2)

    def map_row_by_position(self, cells: List[str]) -> Dict[str, object]:
        trimmed = list(cells)

        if trimmed and "zaznacz tylko ten wiersz" in self.normalize_key(trimmed[0]):
            trimmed = trimmed[1:]

        if len(trimmed) < 10:
            return {}

        identifier = trimmed[0] if len(trimmed) > 0 else ""
        seller_name = trimmed[1] if len(trimmed) > 1 else ""
        nr_ksef = self.clean_ksef_number(trimmed[2]) if len(trimmed) > 2 else ""
        invoice_no = self.clean_invoice_number(trimmed[3]) if len(trimmed) > 3 else ""
        issue_date = self.normalize_export_date(trimmed[4]) if len(trimmed) > 4 else ""
        saved_date = self.normalize_export_date(trimmed[5]) if len(trimmed) > 5 else ""
        received_date = self.normalize_export_date(trimmed[6]) if len(trimmed) > 6 else ""

        netto, waluta = self.parse_amount_and_currency(trimmed[7]) if len(trimmed) > 7 else ("", "")
        brutto, waluta_brutto = self.parse_amount_and_currency(trimmed[8]) if len(trimmed) > 8 else ("", "")
        vat_pln, _ = self.parse_amount_and_currency(trimmed[9]) if len(trimmed) > 9 else ("", "")

        if not waluta:
            waluta = waluta_brutto

        return {
            "Identyfikator sprzedawcy": identifier,
            "Nazwa sprzedawcy": seller_name,
            "Nr KSeF": nr_ksef,
            "Nr faktury": invoice_no,
            "Data wystawienia": issue_date,
            "Data zapisania w KSeF": saved_date,
            "Data otrzymania": received_date,
            "Waluta": waluta,
            "Netto": netto,
            "Brutto": brutto,
            "VAT (PLN)": vat_pln,
        }

    def map_to_target_rows(self, headers: List[str], rows: List[Dict]) -> List[Dict[str, object]]:
        raw_headers = self.build_raw_headers(headers, rows)
        canonical = [self.canonical_header(h) for h in raw_headers]
        mapped = []

        for row in rows:
            cells = row["cells"]

            looks_like_current_ksef_grid = (
                len(cells) >= 10
                and (
                    (cells and "zaznacz tylko ten wiersz" in self.normalize_key(cells[0]))
                    or any("kopiuj numer ksef" in self.normalize_key(c) for c in cells[:5])
                )
            )

            if looks_like_current_ksef_grid:
                normalized = self.map_row_by_position(cells)
                if normalized:
                    mapped.append(normalized)
                    continue

            row_dict = {}
            for idx, header in enumerate(canonical):
                row_dict[header] = cells[idx] if idx < len(cells) else ""

            normalized = {col: row_dict.get(col, "") for col in TARGET_COLUMNS}

            if not normalized["Nr KSeF"] and len(cells) >= 10:
                positional = self.map_row_by_position(cells)
                if positional:
                    normalized = positional

            mapped.append(normalized)

        return mapped

    def to_number_if_possible(self, value: str):
        text = "" if value is None else str(value).strip().replace(" ", "")
        if not text:
            return ""
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")
        if re.fullmatch(r"-?\d+(\.\d+)?", text):
            try:
                return float(text) if "." in text else int(text)
            except ValueError:
                return value
        return value

    def build_summary_rows(self, row_count: int, page_count: int) -> List[Tuple[str, object]]:
        date_from, date_to = self.validate_date_range()
        return [
            ("Data od (opis pliku)", date_from),
            ("Data do (opis pliku)", date_to),
            ("Liczba stron", page_count),
            ("Liczba wierszy", row_count),
            ("Data eksportu", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

    def style_header_row(self, ws, row_no: int = 1):
        fill = PatternFill("solid", fgColor="C81F25")
        font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
        for cell in ws[row_no]:
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def auto_fit_worksheet(self, ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 34)

    def save_excel(self, headers: List[str], rows: List[Dict]) -> str:
        if not rows:
            raise ValueError("Brak danych do zapisania.")

        date_from, date_to = self.validate_date_range()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"KSeF_zestawienie_{date_from}_{date_to}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()
        summary_ws = wb.active
        summary_ws.title = "Podsumowanie"
        summary_ws.append(["Parametr", "Wartość"])
        for row in self.build_summary_rows(len(rows), self.last_pages):
            summary_ws.append(row)
        self.style_header_row(summary_ws)
        self.auto_fit_worksheet(summary_ws)
        summary_ws.freeze_panes = "A2"
        summary_ws.auto_filter.ref = summary_ws.dimensions

        ws = wb.create_sheet("Faktury")
        ws.append(TARGET_COLUMNS)

        for row in self.map_to_target_rows(headers, rows):
            ws.append([
                row["Identyfikator sprzedawcy"],
                row["Nazwa sprzedawcy"],
                row["Nr KSeF"],
                row["Nr faktury"],
                row["Data wystawienia"],
                row["Data zapisania w KSeF"],
                row["Data otrzymania"],
                row["Waluta"],
                row["Netto"],
                row["Brutto"],
                row["VAT (PLN)"],
            ])

        self.style_header_row(ws)
        self.auto_fit_worksheet(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(filepath)
        return filepath

    def run_full_export(self):
        try:
            self.ensure_browser_ready()
            self.validate_date_range()

            self.start_busy("Przygotowanie")
            self.log("[INFO] Start zestawienia.")
            self.log("[INFO] Próbuję zastosować zakres dat w KSeF.")
            self.try_apply_dates()
            if not self.wait_for_table_ready():
                raise RuntimeError("Nie udało się znaleźć tabeli z fakturami. Otwórz listę faktur zakupu w KSeF i spróbuj ponownie.")
            self.log("[INFO] Program pobiera to, co aktualnie widać na liście w KSeF.")
            self.log("[INFO] Skanuję wszystkie strony listy...")
            headers, rows, pages = self.scan_all_pages()
            self.last_headers = headers
            self.last_rows = rows
            self.last_pages = pages

            if not rows:
                self.reset_progress("Brak danych")
                self.summary_var.set("Nie znaleziono żadnych wierszy")
                messagebox.showwarning("Brak danych", "Nie znaleziono żadnych wierszy na liście.")
                return

            self.update_progress(2, 3, "Zapisywanie Excel")
            filepath = self.save_excel(headers, rows)
            self.update_progress(3, 3, "Gotowe")

            self.summary_var.set(f"Znaleziono {len(rows)} wierszy na {pages} stronach")
            self.file_var.set(Path(filepath).name)
            self.log(f"[OK] Zapisano plik: {filepath}")
            messagebox.showinfo("Gotowe", f"Zestawienie zapisane.\n\n{filepath}")
        except Exception as e:
            self.reset_progress("Błąd")
            self.log(f"[BŁĄD] {e}")
            messagebox.showerror("Błąd", str(e))

    def on_close(self):
        self.close_browser_resources()
        self.root.destroy()


if __name__ == "__main__":
    if sys.platform.startswith("win"):
        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

    root = tk.Tk()
    app = KsefSimpleSummaryApp(root)
    root.mainloop()
