import json
import re
import threading
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver import EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

APP_TITLE = "KSeF - Zestawienie faktur zakupu"
KSEF_URL = "https://ksef.mf.gov.pl/"
DEFAULT_SAVE_DIR = str(Path.home() / "Desktop")
CONFIG_PATH = Path(__file__).with_name("ksef_config.json")

TARGET_COLUMNS = [
    "Identyfikator sprzedawcy",
    "Nazwa sprzedawcy",
    "Nr KSeF",
    "Nr faktury",
    "Data wystawienia",
    "Data zapisania w KSeF",
    "Data otrzymania",
    "Waluta",
    "Netto",
    "Brutto",
    "VAT (PLN)",
]

HEADER_ALIASES = {
    "Identyfikator sprzedawcy": [
        "identyfikator sprzedawcy", "nip sprzedawcy", "nip", "sprzedawca nip"
    ],
    "Nazwa sprzedawcy": [
        "nazwa sprzedawcy", "sprzedawca", "nazwa kontrahenta", "kontrahent"
    ],
    "Nr KSeF": [
        "nr ksef", "numer ksef", "ksef", "identyfikator ksef"
    ],
    "Nr faktury": [
        "nr faktury", "numer faktury", "faktura", "numer dokumentu"
    ],
    "Data wystawienia": [
        "data wystawienia", "wystawiono"
    ],
    "Data zapisania w KSeF": [
        "data zapisania w ksef", "zapisano w ksef", "data nadania w ksef", "data przyjęcia w ksef"
    ],
    "Data otrzymania": [
        "data otrzymania", "otrzymano", "data odbioru"
    ],
    "Waluta": [
        "waluta", "currency"
    ],
    "Netto": [
        "netto", "wartość netto", "kwota netto"
    ],
    "Brutto": [
        "brutto", "wartość brutto", "kwota brutto"
    ],
    "VAT (PLN)": [
        "vat (pln)", "vat pln", "kwota vat pln", "vat"
    ],
}

DEFAULT_CONFIG = {
    "table_selectors": [
        "table",
        "[role='grid']",
        "[role='table']",
        ".ag-root-wrapper",
        ".p-datatable",
    ],
    "next_button_xpaths": [
        "//button[contains(., 'Następna') and not(@disabled)]",
        "//button[contains(., 'Dalej') and not(@disabled)]",
        "//button[@aria-label='Next page' and not(@disabled)]",
        "//a[contains(., 'Następna') and not(contains(@class, 'disabled'))]",
        "//button[contains(., '>') and not(@disabled)]",
    ],
    "date_input_hints": {
        "from": [
            "data od", "od", "data otrzymania od", "zakres od", "date from"
        ],
        "to": [
            "data do", "do", "data otrzymania do", "zakres do", "date to"
        ],
    },
    "prefer_date_type": "data_otrzymania",
    "max_pages": 500,
    "wait_timeout": 20,
}


def load_config() -> Dict[str, Any]:
    if CONFIG_PATH.exists():
        try:
            with CONFIG_PATH.open("r", encoding="utf-8") as f:
                loaded = json.load(f)
            merged = DEFAULT_CONFIG.copy()
            merged.update(loaded)
            return merged
        except Exception:
            pass
    with CONFIG_PATH.open("w", encoding="utf-8") as f:
        json.dump(DEFAULT_CONFIG, f, ensure_ascii=False, indent=2)
    return DEFAULT_CONFIG.copy()


def normalize_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("ł", "l").replace("ą", "a").replace("ć", "c").replace("ę", "e")
    text = text.replace("ń", "n").replace("ó", "o").replace("ś", "s").replace("ż", "z").replace("ź", "z")
    text = re.sub(r"\s+", " ", text)
    return text


def parse_amount(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\xa0", " ").replace(" ", "")
    text = text.replace("PLN", "").replace("EUR", "")
    text = text.replace(",", ".")
    try:
        return float(text)
    except Exception:
        return None


def parse_date(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%Y.%m.%d", "%d/%m/%Y"]:
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return text


@dataclass
class RowData:
    values: Dict[str, Any]


class KSeFExporter:
    def __init__(self, logger):
        self.logger = logger
        self.config = load_config()
        self.driver: Optional[webdriver.Edge] = None
        self.wait: Optional[WebDriverWait] = None

    def open_browser(self) -> None:
        if self.driver:
            try:
                self.driver.current_url
                self.driver.maximize_window()
                self.driver.get(KSEF_URL)
                self.logger("Przeglądarka już była otwarta. Otworzyłem stronę KSeF.")
                return
            except Exception:
                self.driver = None
                self.wait = None

        self.logger("Uruchamianie Microsoft Edge...")
        options = EdgeOptions()
        options.use_chromium = True
        options.add_argument("--start-maximized")
        options.add_argument("--disable-notifications")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)

        self.driver = webdriver.Edge(options=options)
        self.wait = WebDriverWait(self.driver, self.config.get("wait_timeout", 20))
        self.driver.get(KSEF_URL)
        self.logger("Otworzyłem KSeF. Zaloguj się ręcznie w przeglądarce.")

    def ensure_browser(self) -> None:
        if not self.driver:
            raise RuntimeError("Najpierw kliknij 'Otwórz KSeF'.")

    def set_date_filters_best_effort(self, date_from: str, date_to: str) -> bool:
        self.ensure_browser()
        self.logger("Próbuję ustawić zakres dat w KSeF...")
        hints = self.config.get("date_input_hints", {})

        script = r"""
        const fromValue = arguments[0];
        const toValue = arguments[1];
        const hints = arguments[2];

        function norm(s) {
          return (s || '')
            .toString()
            .toLowerCase()
            .normalize('NFD').replace(/[\u0300-\u036f]/g, '')
            .replace(/\s+/g, ' ')
            .trim();
        }

        function collectCandidates() {
          const items = [];
          const inputs = Array.from(document.querySelectorAll('input, textarea'));
          for (const el of inputs) {
            const placeholder = norm(el.getAttribute('placeholder'));
            const aria = norm(el.getAttribute('aria-label'));
            const name = norm(el.getAttribute('name'));
            const id = el.getAttribute('id');
            let labelText = '';
            if (id) {
              const lbl = document.querySelector(`label[for="${id}"]`);
              if (lbl) labelText += ' ' + lbl.innerText;
            }
            const parent = el.closest('div, label, section, article, td, th');
            if (parent) labelText += ' ' + parent.innerText;
            items.push({
              el,
              text: norm([placeholder, aria, name, labelText].join(' ')),
              type: (el.getAttribute('type') || '').toLowerCase()
            });
          }
          return items;
        }

        function matchInput(candidates, hintList) {
          for (const hint of hintList) {
            const nh = norm(hint);
            const matched = candidates.find(c => c.text.includes(nh));
            if (matched) return matched.el;
          }
          return null;
        }

        function setValue(el, value) {
          if (!el) return false;
          el.focus();
          el.value = value;
          el.dispatchEvent(new Event('input', { bubbles: true }));
          el.dispatchEvent(new Event('change', { bubbles: true }));
          el.dispatchEvent(new KeyboardEvent('keydown', { bubbles: true, key: 'Enter' }));
          el.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: 'Enter' }));
          return true;
        }

        const candidates = collectCandidates();
        const fromEl = matchInput(candidates, hints.from || []);
        const toEl = matchInput(candidates, hints.to || []);
        const okFrom = setValue(fromEl, fromValue);
        const okTo = setValue(toEl, toValue);
        return { okFrom, okTo, candidates: candidates.map(c => c.text).slice(0, 30) };
        """
        result = self.driver.execute_script(script, date_from, date_to, hints)
        ok = bool(result and result.get("okFrom") and result.get("okTo"))
        if ok:
            self.logger("Daty zostały wpisane. Kliknij w KSeF wyszukiwanie/filtr, jeśli portal tego wymaga.")
        else:
            self.logger("Nie udało się pewnie wpisać dat automatycznie. Ustaw je ręcznie w KSeF.")
        return ok

    def extract_table_rows(self) -> Tuple[List[RowData], List[str]]:
        self.ensure_browser()
        self.logger("Pobieram dane z aktualnej tabeli...")
        script = r"""
        function visibleText(el) {
          return (el && el.innerText ? el.innerText : '').replace(/\s+/g, ' ').trim();
        }

        function extractFromHtmlTable(table) {
          const headerCells = Array.from(table.querySelectorAll('thead th')).map(visibleText).filter(Boolean);
          let headers = headerCells;
          let rows = [];
          const bodyRows = Array.from(table.querySelectorAll('tbody tr'));
          for (const tr of bodyRows) {
            const cells = Array.from(tr.querySelectorAll('td, th')).map(visibleText);
            if (cells.some(Boolean)) rows.push(cells);
          }
          if (!headers.length && rows.length > 0) {
            headers = rows[0];
            rows = rows.slice(1);
          }
          return { headers, rows };
        }

        function extractFromGrid(grid) {
          let headers = [];
          let rows = [];

          const headerRow = grid.querySelector('[role="row"]');
          if (headerRow) {
            headers = Array.from(headerRow.querySelectorAll('[role="columnheader"], th')).map(visibleText).filter(Boolean);
          }

          const rowEls = Array.from(grid.querySelectorAll('[role="row"]'));
          for (const rowEl of rowEls) {
            const cells = Array.from(rowEl.querySelectorAll('[role="gridcell"], td')).map(visibleText);
            if (cells.some(Boolean)) rows.push(cells);
          }

          if (!headers.length && rows.length) {
            headers = rows[0];
            rows = rows.slice(1);
          }
          return { headers, rows };
        }

        const selectors = arguments[0] || [];
        let best = { headers: [], rows: [] };

        for (const selector of selectors) {
          const els = Array.from(document.querySelectorAll(selector));
          for (const el of els) {
            let result = { headers: [], rows: [] };
            if (el.tagName.toLowerCase() === 'table') {
              result = extractFromHtmlTable(el);
            } else {
              result = extractFromGrid(el);
            }
            if (result.rows.length > best.rows.length) {
              best = result;
            }
          }
        }

        if (!best.rows.length) {
          const tables = Array.from(document.querySelectorAll('table'));
          for (const t of tables) {
            const result = extractFromHtmlTable(t);
            if (result.rows.length > best.rows.length) best = result;
          }
        }

        return best;
        """
        result = self.driver.execute_script(script, self.config.get("table_selectors", []))
        headers = result.get("headers") or []
        rows = result.get("rows") or []
        data = [RowData(dict(enumerate(r))) for r in rows if any(str(x).strip() for x in r)]
        return data, headers

    def _map_headers(self, headers: List[str]) -> Dict[int, str]:
        mapping: Dict[int, str] = {}
        norm_headers = [normalize_text(h) for h in headers]
        for index, header in enumerate(norm_headers):
            for target, aliases in HEADER_ALIASES.items():
                if any(alias in header for alias in aliases):
                    mapping[index] = target
                    break

        # fallback by position if table seems very similar
        if len(headers) >= len(TARGET_COLUMNS) and len(mapping) < 5:
            for idx, target in enumerate(TARGET_COLUMNS):
                mapping[idx] = target
        return mapping

    def _normalize_row(self, headers: List[str], row_values: Dict[int, Any]) -> Dict[str, Any]:
        mapped = {col: "" for col in TARGET_COLUMNS}
        mapping = self._map_headers(headers)
        for index, raw_value in row_values.items():
            target = mapping.get(index)
            if target:
                mapped[target] = raw_value

        # lekkie czyszczenie
        for col in ["Data wystawienia", "Data zapisania w KSeF", "Data otrzymania"]:
            mapped[col] = parse_date(mapped[col])
        for col in ["Netto", "Brutto", "VAT (PLN)"]:
            amount = parse_amount(mapped[col])
            mapped[col] = amount if amount is not None else mapped[col]
        mapped["Waluta"] = str(mapped["Waluta"] or "").strip().upper()
        mapped["Nr KSeF"] = str(mapped["Nr KSeF"] or "").strip()
        mapped["Nr faktury"] = str(mapped["Nr faktury"] or "").strip()
        mapped["Nazwa sprzedawcy"] = str(mapped["Nazwa sprzedawcy"] or "").strip()
        mapped["Identyfikator sprzedawcy"] = str(mapped["Identyfikator sprzedawcy"] or "").strip()
        return mapped

    def click_next_page(self) -> bool:
        self.ensure_browser()
        xpaths = self.config.get("next_button_xpaths", [])
        for xp in xpaths:
            try:
                btn = self.driver.find_element(By.XPATH, xp)
                if btn.is_displayed() and btn.is_enabled():
                    self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                    btn.click()
                    return True
            except Exception:
                continue
        return False

    def collect_all_pages(self) -> List[Dict[str, Any]]:
        self.ensure_browser()
        seen_signatures = set()
        collected: List[Dict[str, Any]] = []
        max_pages = int(self.config.get("max_pages", 500))

        for page_idx in range(1, max_pages + 1):
            rows, headers = self.extract_table_rows()
            if not rows:
                if page_idx == 1:
                    raise RuntimeError(
                        "Nie znalazłem tabeli z danymi. Wejdź w KSeF na listę faktur zakupu i spróbuj ponownie."
                    )
                break

            normalized_rows = [self._normalize_row(headers, row.values) for row in rows]
            signature = tuple(tuple(r.get(c, "") for c in TARGET_COLUMNS) for r in normalized_rows[:10])
            if signature in seen_signatures:
                self.logger("Wykryłem powtórzenie strony - kończę pobieranie.")
                break
            seen_signatures.add(signature)

            before = len(collected)
            for row in normalized_rows:
                if any(str(row.get(c, "")).strip() for c in TARGET_COLUMNS):
                    collected.append(row)
            added = len(collected) - before
            self.logger(f"Strona {page_idx}: dodano {added} wierszy.")

            if not self.click_next_page():
                self.logger("To była ostatnia strona.")
                break

            WebDriverWait(self.driver, self.config.get("wait_timeout", 20)).until(
                lambda d: True
            )
            self.driver.implicitly_wait(1)
        return self._deduplicate(collected)

    @staticmethod
    def _deduplicate(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        seen = set()
        result = []
        for row in rows:
            key = tuple(str(row.get(c, "")).strip() for c in TARGET_COLUMNS)
            if key in seen:
                continue
            seen.add(key)
            result.append(row)
        return result

    def save_to_excel(self, rows: List[Dict[str, Any]], output_path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Faktury zakupu"

        ws.append(TARGET_COLUMNS)
        for row in rows:
            ws.append([row.get(c, "") for c in TARGET_COLUMNS])

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        currency_columns = {
            "Netto": 9,
            "Brutto": 10,
            "VAT (PLN)": 11,
        }
        for _, col_idx in currency_columns.items():
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col_idx).number_format = '#,##0.00'

        widths = {
            "A": 20,
            "B": 42,
            "C": 34,
            "D": 24,
            "E": 18,
            "F": 22,
            "G": 18,
            "H": 10,
            "I": 14,
            "J": 14,
            "K": 14,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        wb.save(output_path)

    def close(self) -> None:
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None
            self.wait = None


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("860x620")
        self.root.minsize(820, 560)

        self.exporter = KSeFExporter(self.log)
        self.save_dir = tk.StringVar(value=DEFAULT_SAVE_DIR)
        self.date_from = tk.StringVar(value=datetime.today().replace(day=1).strftime("%Y-%m-%d"))
        self.date_to = tk.StringVar(value=datetime.today().strftime("%Y-%m-%d"))
        self.open_after_save = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="Gotowy")

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def _build_ui(self) -> None:
        style = ttk.Style()
        try:
            style.theme_use("vista")
        except Exception:
            pass

        container = ttk.Frame(self.root, padding=16)
        container.pack(fill="both", expand=True)

        title = ttk.Label(container, text=APP_TITLE, font=("Segoe UI", 16, "bold"))
        title.pack(anchor="w")

        subtitle = ttk.Label(
            container,
            text="1. Otwórz KSeF  2. Zaloguj się ręcznie  3. Ustaw daty lub kliknij pobieranie  4. Zapisz Excel",
            font=("Segoe UI", 10),
        )
        subtitle.pack(anchor="w", pady=(4, 12))

        card = ttk.LabelFrame(container, text="Ustawienia", padding=12)
        card.pack(fill="x")

        row1 = ttk.Frame(card)
        row1.pack(fill="x", pady=4)
        ttk.Label(row1, text="Data od (YYYY-MM-DD):", width=22).pack(side="left")
        ttk.Entry(row1, textvariable=self.date_from, width=18).pack(side="left", padx=(0, 12))
        ttk.Label(row1, text="Data do (YYYY-MM-DD):", width=20).pack(side="left")
        ttk.Entry(row1, textvariable=self.date_to, width=18).pack(side="left")

        row2 = ttk.Frame(card)
        row2.pack(fill="x", pady=8)
        ttk.Label(row2, text="Folder zapisu:", width=22).pack(side="left")
        ttk.Entry(row2, textvariable=self.save_dir, width=60).pack(side="left", padx=(0, 8), fill="x", expand=True)
        ttk.Button(row2, text="Wybierz...", command=self.choose_folder).pack(side="left")

        row3 = ttk.Frame(card)
        row3.pack(fill="x", pady=(8, 0))
        ttk.Checkbutton(row3, text="Otwórz plik po zapisaniu", variable=self.open_after_save).pack(side="left")

        actions = ttk.LabelFrame(container, text="Akcje", padding=12)
        actions.pack(fill="x", pady=12)

        btn_row = ttk.Frame(actions)
        btn_row.pack(fill="x")
        ttk.Button(btn_row, text="1. Otwórz KSeF", command=self.open_ksef).pack(side="left", padx=(0, 8))
        ttk.Button(btn_row, text="2. Ustaw daty w KSeF", command=self.set_dates).pack(side="left", padx=(0, 8))
        ttk.Button(btn_row, text="3. Pobierz zestawienie", command=self.start_export).pack(side="left")

        hint = ttk.Label(
            actions,
            text=(
                "Jeśli automatyczne wpisanie dat nie zadziała, ustaw daty ręcznie w KSeF, "
                "otwórz listę faktur zakupu i kliknij 'Pobierz zestawienie'."
            ),
            wraplength=780,
            foreground="#555555",
        )
        hint.pack(anchor="w", pady=(10, 0))

        log_frame = ttk.LabelFrame(container, text="Log", padding=8)
        log_frame.pack(fill="both", expand=True)

        self.log_box = tk.Text(log_frame, height=20, wrap="word", font=("Consolas", 10))
        self.log_box.pack(side="left", fill="both", expand=True)
        scroll = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_box.yview)
        scroll.pack(side="right", fill="y")
        self.log_box.configure(yscrollcommand=scroll.set)

        status_bar = ttk.Label(container, textvariable=self.status_var, anchor="w")
        status_bar.pack(fill="x", pady=(8, 0))

        self.log("Program gotowy.")
        self.log("Kliknij 'Otwórz KSeF', zaloguj się ręcznie i przejdź do listy faktur zakupu.")

    def choose_folder(self) -> None:
        folder = filedialog.askdirectory(initialdir=self.save_dir.get() or DEFAULT_SAVE_DIR)
        if folder:
            self.save_dir.set(folder)

    def log(self, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_box.insert("end", f"[{timestamp}] {message}\n")
        self.log_box.see("end")
        self.status_var.set(message)
        self.root.update_idletasks()

    def validate_dates(self) -> Tuple[str, str]:
        date_from = self.date_from.get().strip()
        date_to = self.date_to.get().strip()
        try:
            df = datetime.strptime(date_from, "%Y-%m-%d")
            dt = datetime.strptime(date_to, "%Y-%m-%d")
        except ValueError:
            raise RuntimeError("Daty muszą mieć format YYYY-MM-DD.")
        if df > dt:
            raise RuntimeError("Data od nie może być późniejsza niż data do.")
        return date_from, date_to

    def open_ksef(self) -> None:
        def worker():
            try:
                self.exporter.open_browser()
            except Exception as e:
                self.log(f"Błąd otwierania KSeF: {e}")
                messagebox.showerror(APP_TITLE, f"Nie udało się otworzyć KSeF.\n\n{e}")
        threading.Thread(target=worker, daemon=True).start()

    def set_dates(self) -> None:
        try:
            date_from, date_to = self.validate_dates()
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))
            return

        def worker():
            try:
                ok = self.exporter.set_date_filters_best_effort(date_from, date_to)
                if ok:
                    self.log("Zakres dat wpisany. Sprawdź w KSeF i uruchom pobieranie.")
                else:
                    self.log("Ustaw daty ręcznie w KSeF i kliknij pobieranie.")
            except Exception as e:
                self.log(f"Błąd ustawiania dat: {e}")
                messagebox.showerror(APP_TITLE, f"Nie udało się ustawić dat.\n\n{e}")
        threading.Thread(target=worker, daemon=True).start()

    def start_export(self) -> None:
        try:
            date_from, date_to = self.validate_dates()
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))
            return

        def worker():
            try:
                self.log("Start pobierania...")
                rows = self.exporter.collect_all_pages()
                if not rows:
                    raise RuntimeError("Nie pobrano żadnych danych.")

                out_dir = Path(self.save_dir.get().strip() or DEFAULT_SAVE_DIR)
                out_dir.mkdir(parents=True, exist_ok=True)
                filename = f"Zestawienie_faktur_zakupu_{date_from}_{date_to}.xlsx"
                output_path = str(out_dir / filename)
                self.exporter.save_to_excel(rows, output_path)
                self.log(f"Gotowe. Zapisano {len(rows)} wierszy do pliku: {output_path}")

                if self.open_after_save.get():
                    try:
                        import os
                        os.startfile(output_path)  # type: ignore[attr-defined]
                    except Exception:
                        pass
                messagebox.showinfo(APP_TITLE, f"Zapisano plik:\n{output_path}\n\nLiczba wierszy: {len(rows)}")
            except Exception as e:
                self.log(f"Błąd: {e}")
                details = traceback.format_exc(limit=3)
                self.log(details)
                messagebox.showerror(APP_TITLE, f"Wystąpił błąd:\n\n{e}")

        threading.Thread(target=worker, daemon=True).start()

    def on_close(self) -> None:
        try:
            self.exporter.close()
        finally:
            self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
